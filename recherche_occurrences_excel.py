#!/usr/bin/env python3
"""
Programme pour rechercher des occurrences dans un fichier Excel 
et créer une liste de liens dans un nouvel onglet.

Utilisation:
    python recherche_occurrences_excel.py <fichier_excel> <terme_recherche> [--onglet_sortie NOM_ONGLET]
    python recherche_occurrences_excel.py <fichier_excel> <terme_recherche> --tous_onglets
"""

import argparse
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.cell import Hyperlink


def rechercher_occurrences(fichier_excel, terme_recherche, onglet_sortie="Résultats", tous_onglets=False):
    """
    Recherche des occurrences d'un terme dans un fichier Excel et crée une liste de liens.
    
    Args:
        fichier_excel: Chemin vers le fichier Excel
        terme_recherche: Terme à rechercher
        onglet_sortie: Nom de l'onglet où écrire les résultats
        tous_onglets: Si True, recherche dans tous les onglets
    """
    try:
        # Charger le classeur
        wb = load_workbook(fichier_excel)
        
        # Créer ou obtenir l'onglet de sortie
        if onglet_sortie in wb.sheetnames:
            del wb[onglet_sortie]
        ws_sortie = wb.create_sheet(title=onglet_sortie)
        
        # Écrire l'en-tête
        ws_sortie.append(["Onglet", "Cellule", "Contenu", "Lien"])
        
        # Définir les styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        for cell in ws_sortie[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        ligne_sortie = 2
        
        # Déterminer les onglets à analyser
        if tous_onglets:
            onglets_a_analyser = wb.sheetnames
        else:
            # Par défaut, analyser l'onglet actif ou le premier
            onglets_a_analyser = [wb.sheetnames[0]]
        
        # Rechercher dans chaque onglet
        for nom_onglet in onglets_a_analyser:
            if nom_onglet == onglet_sortie:
                continue
                
            ws = wb[nom_onglet]
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        valeur_str = str(cell.value)
                        if terme_recherche.lower() in valeur_str.lower():
                            # Créer un lien vers la cellule
                            lien = f"#{nom_onglet}!{cell.coordinate}"
                            
                            # Écrire les informations dans l'onglet de sortie
                            ws_sortie.append([
                                nom_onglet,
                                cell.coordinate,
                                valeur_str[:100] + "..." if len(valeur_str) > 100 else valeur_str,
                                "Lien"
                            ])
                            
                            # Ajouter l'hyperlien à la dernière cellule
                            last_cell = ws_sortie.cell(row=ligne_sortie, column=4)
                            last_cell.hyperlink = Hyperlink(ref=cell.coordinate, location=lien)
                            last_cell.font = Font(color="0000FF", underline="single")
                            
                            ligne_sortie += 1
        
        # Sauvegarder le fichier
        wb.save(fichier_excel)
        print(f"Recherche terminée. {ligne_sortie - 2} occurrences trouvées.")
        print(f"Les résultats ont été enregistrés dans l'onglet '{onglet_sortie}'.")
        
    except FileNotFoundError:
        print(f"Erreur: Le fichier '{fichier_excel}' est introuvable.")
        sys.exit(1)
    except Exception as e:
        print(f"Une erreur s'est produite: {str(e)}")
        sys.exit(1)


def main():
    parser = argparse.ArgumentParser(
        description="Rechercher des occurrences dans un fichier Excel et créer une liste de liens."
    )
    parser.add_argument("fichier_excel", help="Chemin vers le fichier Excel")
    parser.add_argument("terme_recherche", help="Terme à rechercher")
    parser.add_argument("--onglet_sortie", default="Résultats", 
                       help="Nom de l'onglet de sortie (par défaut: Résultats)")
    parser.add_argument("--tous_onglets", action="store_true",
                       help="Rechercher dans tous les onglets")
    
    args = parser.parse_args()
    
    rechercher_occurrences(
        args.fichier_excel,
        args.terme_recherche,
        args.onglet_sortie,
        args.tous_onglets
    )


if __name__ == "__main__":
    main()
